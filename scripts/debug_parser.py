import openpyxl
import sys
import os

# Add app to sys.path
sys.path.append(os.getcwd())

from app.services.excel_parser import (
    _normalizar, _normalize_date, _should_discard, SKIP_SHEETS
)

# Usa cualquiera de los archivos migrados que tengas disponible
ARCHIVO = "/Users/tonirosascastillo/Code/limpios/Grupos Aptitud 8 9 PM.xlsx"
HOJA = "Grupo Caminata 7"

print(f"Abriendo archivo: {ARCHIVO}")
wb = openpyxl.load_workbook(ARCHIVO, data_only=True)
ws = wb[HOJA]

# 1. Detección de formato
fila1_col1 = _normalizar(ws.cell(row=1, column=1).value)
fila2_col1 = _normalizar(ws.cell(row=2, column=1).value)
print(f"fila1_col1_norm: {repr(fila1_col1)}")
print(f"fila2_col1_norm: {repr(fila2_col1)}")

if fila2_col1 in ("nombre", "name", "folio"):
    header_row_idx = 1
    data_start_row_idx = 2
elif fila1_col1 in ("nombre", "name", "folio"):
    header_row_idx = 0
    data_start_row_idx = 1
else:
    print(f"No se detectaron headers en fila 1 ni 2.")
    header_row_idx = None

# 2. Headers de la fila que se usaría como header_row
rows = list(ws.iter_rows(values_only=True))
if header_row_idx is not None:
    header_row = rows[header_row_idx]
    print(f"\n--- Headers detectados (fila {header_row_idx+1}) ---")
    for i, h in enumerate(header_row[:15]):
        norm = _normalizar(h)
        fecha = _normalize_date(h)
        discard = _should_discard(str(h or ""))
        print(f"  col{i+1}: valor={repr(h)} | norm={repr(norm)} | fecha={fecha} | discard={discard}")
else:
    header_row = []

# 3. Verificar has_meta y has_dates
from app.services.excel_parser import META_COLS, SUMMARY_COLS
print(f"\n--- Matching meta/date columns ---")
has_meta = False
has_dates = False
for i, h in enumerate(header_row):
    if h is None:
        continue
    h_norm = _normalizar(h)
    for mc in META_COLS:
        if _normalizar(mc) == h_norm:
            print(f"  META match: col{i+1} '{h}' -> '{mc}'")
            has_meta = True
    parsed = _normalize_date(h)
    if parsed:
        print(f"  DATE match: col{i+1} '{h}' -> {parsed}")
        has_dates = True

print(f"\nhas_meta={has_meta}, has_dates={has_dates}")
print(f"Resultado esperado: ambos True para parsear alumnos")

# 4. Primer alumno
print(f"\n--- Fila 3 (primer alumno) ---")
alumno_row = rows[2]
for i, v in enumerate(alumno_row[:8]):
    print(f"  col{i+1}: {repr(v)} | tipo={type(v).__name__}")
