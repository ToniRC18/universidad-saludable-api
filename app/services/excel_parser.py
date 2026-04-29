"""
Excel parser for Universidad Saludable attendance sheets.

Expected structure per sheet:
  Row 0  → visual header (schedule + "Semana 1"..."Semana 12") — IGNORED
  Row 1  → real column names
  Rows 2+→ one student per row
"""
from __future__ import annotations

import logging
import re
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def leer_meta(wb: openpyxl.Workbook) -> dict:
    """Busca la hoja _meta en el workbook y extrae semestre_id e horario_id."""
    if "_meta" not in wb.sheetnames:
        raise ValueError("Falta la hoja _meta. Descarga la plantilla oficial desde el módulo de semestres.")
    
    ws = wb["_meta"]
    meta_data = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) >= 2 and row[0] is not None:
            key = str(row[0]).strip()
            val = row[1]
            meta_data[key] = val
            
    try:
        semestre_id = int(meta_data.get("semestre_id"))
        horario_id = int(meta_data.get("horario_id"))
    except (TypeError, ValueError):
        raise ValueError("El archivo no es una plantilla válida del sistema. Descarga la plantilla desde el módulo de semestres.")
        
    return {"semestre_id": semestre_id, "horario_id": horario_id}


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Columns to DISCARD (case-insensitive substring match on header name)
# Album Entregado y Cuestionario se descartan si existen; si no están presentes, no ocurre nada.
DISCARD_PATTERNS = [
    r"^$",                          # blank / empty header
    r"telefono",
    r"cuestionario",                # optional column — descartada si existe
    r"album\s*entregado",           # optional column — descartada si existe
    r"e42",
    r"inhabil|inh[aá]bil|asueto",   # días inhábiles (cualquier variante con/sin tilde)
]

# Columns we ALWAYS want to keep (exact names used in the Excel)
SUMMARY_COLS = {
    "ASISTENCIA",
    "NUTRICIÓN",
    "FISIO",
    "LIMPIEZA",
    "COAE",
    "TALLER",
    "TOTAL",
}

# Metadata columns (exact names)
META_COLS = {"Folio", "Nombre", "Matricula", "Semestre", "Carrera"}

# Month name → month number  (Spanish abbreviations used in the Excel)
MONTH_MAP = {
    "ene": 1, "enero": 1,
    "feb": 2, "febrero": 2,
    "mar": 3, "marzo": 3,
    "abr": 4, "abril": 4,
    "may": 5, "mayo": 5,
    "jun": 6, "junio": 6,
    "jul": 7, "julio": 7,
    "ago": 8, "agosto": 8,
    "sep": 9, "sept": 9, "septiembre": 9,
    "oct": 10, "octubre": 10,
    "nov": 11, "noviembre": 11,
    "dic": 12, "diciembre": 12,
    # Common abbreviations that appear in the Excel
    "marz": 3, "abrl": 4, "agos": 8,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

import unicodedata

def _normalizar(s):
    """Normaliza texto removiendo espacios, pasando a minúsculas, quitando tildes y normalizando unicode."""
    if s is None:
        return ""
    # Normalización unicode para manejar acentos compuestos (e.g. i + ´ vs í)
    s = unicodedata.normalize('NFKD', str(s))
    return (s.strip().lower()
            .replace("á","a").replace("é","e")
            .replace("í","i").replace("ó","o").replace("ú","u")
            .replace("\u0301", "")) # Eliminar diacríticos residuales de NFKD


def _should_discard(header: str) -> bool:
    """Return True if this column header should be discarded."""
    h = _normalizar(header)
    for pattern in DISCARD_PATTERNS:
        if re.search(pattern, h, re.IGNORECASE):
            return True
    return False


def _normalize_date(value: Any, year_hint: int = datetime.now().year) -> date | None:
    """
    Try to convert *value* to a Python date.

    Handles:
    - datetime objects (from openpyxl)
    - date objects
    - Strings like "3 Marz", "17 Febrero", "3/3", "2025-03-03"
    """
    if value is None:
        return None

    # NUEVO: formato DD/MM/YYYY (plantillas nuevas)
    if isinstance(value, str) and '/' in value:
        parts = value.strip().split('/')
        if len(parts) == 3:
            try:
                return date(int(parts[2]), int(parts[1]), int(parts[0]))
            except:
                pass

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text or text.lower() in ("nan", "none", ""):
        return None

    # ISO format
    try:
        if '-' in text and len(text) >= 10:
            return date.fromisoformat(text[:10])
    except ValueError:
        pass

    # "DD/MM" or "DD/MM/YYYY" (plantillas nuevas y manuales)
    slash_match = re.match(r"^(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?$", text)
    if slash_match:
        d, m, y = slash_match.groups()
        y = int(y) if y else year_hint
        try:
            d_int, m_int, y_int = int(d), int(m), int(y)
            if y_int < 100:
                y_int += 2000
            return date(y_int, m_int, d_int)
        except ValueError:
            pass

    # "DD MonthName" or "DD MonthAbbrev" — e.g. "3 Marz", "17 Febrero"
    text_match = re.match(r"^(\d{1,2})\s+([A-Za-záéíóúñÁÉÍÓÚÑ]+)", text)
    if text_match:
        day_str, month_str = text_match.groups()
        month_num = MONTH_MAP.get(month_str.lower().rstrip("."))
        if month_num:
            try:
                return date(year_hint, month_num, int(day_str))
            except ValueError:
                pass

    # No loguear warning para headers conocidos que no son fechas (Nombre, Matrícula, etc)
    h_norm = _normalizar(value)
    if h_norm in ["nombre", "matricula", "carrera", "folio", "semestre"]:
        return None
        
    # logger.debug("Could not parse date value: %r", value) # Silenciamos totalmente para evitar ruido
    return None


def _is_date_column(header: str) -> bool:
    """
    Heuristic: a column is a 'date column' if its header can be parsed as a date
    OR looks like a date-ish string ("3 Marz", "17 Febrero", datetime repr, etc.).
    """
    if isinstance(header, (datetime, date)):
        return True
    text = str(header).strip()
    # Already parseable as date?
    if _normalize_date(text) is not None:
        return True
    return False


# (Helper removed, logic moved inside parse_excel loop as per request)


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

class ParsedGrupo:
    def __init__(self, nombre: str, horario: str):
        self.nombre = nombre
        self.horario = horario
        self.max_asistencia: float = 0.0   # sesiones_reales * 2.5, calculado al parsear
        self.alumnos: list[dict] = []      # list of {meta: ..., dates: ..., summary: ...}


# System sheets to skip
SKIP_SHEETS = {"_catalogo", "_meta"}


def parse_excel(file_bytes: bytes, semestre_label: str = "") -> list[ParsedGrupo]:
    """
    Parse an Excel workbook and return a list of ParsedGrupo objects.
    Sheets with unexpected structure are skipped (logged as warnings).
    """
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    grupos: list[ParsedGrupo] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.title in SKIP_SHEETS:
            continue
            
        # Detectar formato por fila (Cambio 1) - USAR _normalizar para robustez
        fila1_col1 = _normalizar(ws.cell(row=1, column=1).value)
        fila2_col1 = _normalizar(ws.cell(row=2, column=1).value)

        if fila2_col1 in ("nombre", "name", "folio"):
            header_row_idx = 1  # 0-indexed para list(rows)
            data_start_row_idx = 2
            horario_row_idx = 0
        elif fila1_col1 in ("nombre", "name", "folio"):
            header_row_idx = 0
            data_start_row_idx = 1
            horario_row_idx = -1
        else:
            logger.warning(f"Sheet '{ws.title}' no tiene headers reconocibles en Col 1 (Fila 1: {repr(fila1_col1)}, Fila 2: {repr(fila2_col1)}) — skipping.")
            continue

        rows = list(ws.iter_rows(values_only=True))
        if header_row_idx >= len(rows):
            continue

        header_row = rows[header_row_idx]    # Real column names
        data_rows = rows[data_start_row_idx:] # Students
        
        # Extract horario text
        horario = ""
        if horario_row_idx != -1 and horario_row_idx < len(rows):
            horario = next((str(c) for c in rows[horario_row_idx] if c is not None), "")

        # Build column index map
        col_map: dict[int, str] = {}      # col_index → role ("meta:<name>", "date:<isodate>", "summary:<name>", "skip")
        date_year = _infer_year(horario, semestre_label)

        for idx, header in enumerate(header_row):
            if header is None:
                col_map[idx] = "skip"
                continue

            # Cambio 3: normalizar header
            h_norm = _normalizar(header)

            if _should_discard(header):
                col_map[idx] = "skip"
                continue

            # Summary columns — comparar normalizando
            matched_summary = None
            for sc in SUMMARY_COLS:
                if _normalizar(sc) == h_norm:
                    matched_summary = sc
                    break
            if matched_summary:
                col_map[idx] = f"summary:{matched_summary}"
                continue

            # Meta columns — comparar normalizando
            matched_meta = None
            for mc in META_COLS:
                if _normalizar(mc) == h_norm:
                    matched_meta = mc
                    break
            if matched_meta:
                col_map[idx] = f"meta:{matched_meta}"
                continue

            # Date columns
            parsed_date = _normalize_date(header, year_hint=date_year)
            if parsed_date is not None:
                col_map[idx] = f"date:{parsed_date.isoformat()}"
                continue

            # Unknown column → skip
            logger.debug("Sheet '%s', col %d ('%s') not recognized — skipping.", sheet_name, idx, str(header))
            col_map[idx] = "skip"

        # Validate we have at least some useful columns
        has_meta = any(v.startswith("meta:") for v in col_map.values())
        has_dates = any(v.startswith("date:") for v in col_map.values())

        if not has_meta or not has_dates:
            logger.warning(
                "Sheet '%s' is missing meta or date columns (meta=%s, dates=%s) — skipping.",
                sheet_name, has_meta, has_dates,
            )
            continue

        # Cambio 1: max_asistencia dinámico — contar sesiones reales (date cols ya filtradas de inhábiles)
        date_col_count = sum(1 for v in col_map.values() if v.startswith("date:"))
        max_asistencia = round(date_col_count * 2.5, 2)

        grupo = ParsedGrupo(nombre=sheet_name, horario=horario)
        grupo.max_asistencia = max_asistencia

        for row in data_rows:
            # Skip completely empty rows
            if not any(c is not None for c in row):
                continue

            alumno: dict = {
                "meta": {},
                "dates": {},   # iso_date_str → numeric value
                "summary": {},
            }

            for idx, role in col_map.items():
                if idx >= len(row):
                    continue
                cell_val = row[idx]

                if role == "skip":
                    continue
                elif role.startswith("meta:"):
                    key = role[5:]
                    alumno["meta"][key.lower()] = str(cell_val).strip() if cell_val is not None else None
                elif role.startswith("date:"):
                    iso = role[5:]
                    alumno["dates"][iso] = _to_numeric(cell_val)
                elif role.startswith("summary:"):
                    key = role[8:]
                    alumno["summary"][key] = _to_numeric(cell_val)

            # Skip rows with no name and no folio (likely empty/junk rows)
            if not alumno["meta"].get("nombre") and not alumno["meta"].get("folio"):
                continue

            grupo.alumnos.append(alumno)

        if not grupo.alumnos:
            logger.warning("Sheet '%s' parsed 0 students — skipping.", sheet_name)
            continue

        grupos.append(grupo)

    return grupos


def _to_numeric(value: Any) -> float:
    """Convert a cell value to float, defaulting to 0."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _infer_year(horario: str, semestre_label: str) -> int:
    """
    Try to extract the year from horario text or semestre_label.
    Defaults to current year.
    """
    for text in (horario, semestre_label):
        m = re.search(r"\b(20\d{2})\b", text)
        if m:
            return int(m.group(1))
    return datetime.now().year
