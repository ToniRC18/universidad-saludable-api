"""
Generación y parseo del Excel plantilla para pruebas físicas.
"""
import logging
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.models.pruebas import PeriodoSeguimiento, PruebaFisica, ResultadoPrueba, SeguimientoGrupo
from app.models import Grupo, Alumno, Upload

logger = logging.getLogger(__name__)

FIXED_COLS = ["Matricula", "Nombre"]

EMPTY_ROWS_FALLBACK = 30


# ---------------------------------------------------------------------------
# Generación
# ---------------------------------------------------------------------------

def generar_plantilla(db: Session, periodo_id: int) -> bytes:
    periodo = db.query(PeriodoSeguimiento).filter(PeriodoSeguimiento.id == periodo_id).first()
    if not periodo:
        raise HTTPException(status_code=404, detail=f"Periodo {periodo_id} no encontrado.")

    seguimiento = periodo.seguimiento
    grupos = seguimiento.grupos
    pruebas = seguimiento.pruebas

    if not grupos:
        raise HTTPException(
            status_code=400,
            detail="El seguimiento no tiene grupos configurados. Agrega al menos un grupo antes de generar la plantilla.",
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    meta_font   = Font(italic=True, color="888888", size=9)

    for grupo in grupos:
        ws = wb.create_sheet(title=grupo.nombre_grupo[:31])

        # ── Fila 0: metadatos ────────────────────────────────────────────
        ws.append([
            seguimiento.nombre,
            periodo.nombre_periodo,
            str(periodo.fecha),
            periodo.semestre_label,
        ])
        for cell in ws[1]:
            cell.font = meta_font

        # ── Fila 1: encabezados ──────────────────────────────────────────
        headers = FIXED_COLS + [p.nombre for p in pruebas]
        ws.append(headers)
        for cell in ws[2]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 16  # Matricula
        ws.column_dimensions["B"].width = 32  # Nombre
        for i, _ in enumerate(pruebas, start=3):
            col_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[col_letter].width = 14

        # ── Filas de datos ───────────────────────────────────────────────
        alumnos = _alumnos_del_grupo(db, grupo)

        if alumnos:
            for alumno in alumnos:
                ws.append([alumno.matricula or "", alumno.nombre or ""] + [""] * len(pruebas))
        else:
            logger.info(
                "Grupo '%s' sin alumnos vinculados — plantilla con %d filas vacías.",
                grupo.nombre_grupo, EMPTY_ROWS_FALLBACK,
            )
            for _ in range(EMPTY_ROWS_FALLBACK):
                ws.append([""] * len(headers))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Alumnos del grupo
# ---------------------------------------------------------------------------

def _alumnos_del_grupo(db: Session, grupo: SeguimientoGrupo) -> list[Alumno]:
    """
    Devuelve los alumnos asignados al grupo del seguimiento.
    Si el grupo fue importado desde una carga (upload_grupo_ref_id != None),
    trae exactamente los alumnos de ese grupo de asistencias.
    Si fue creado manualmente, devuelve lista vacía (filas vacías en plantilla).
    """
    if grupo.upload_grupo_ref_id is None:
        return []

    return (
        db.query(Alumno)
        .filter(Alumno.grupo_id == grupo.upload_grupo_ref_id)
        .order_by(Alumno.nombre)
        .all()
    )


# ---------------------------------------------------------------------------
# Parseo
# ---------------------------------------------------------------------------

def parsear_resultados(
    db: Session,
    periodo_id: int,
    file_bytes: bytes,
) -> dict:
    periodo = db.query(PeriodoSeguimiento).filter(PeriodoSeguimiento.id == periodo_id).first()
    if not periodo:
        raise HTTPException(status_code=404, detail=f"Periodo {periodo_id} no encontrado.")

    seguimiento = periodo.seguimiento
    pruebas: list[PruebaFisica] = seguimiento.pruebas
    grupos: list[SeguimientoGrupo] = seguimiento.grupos

    # índice de grupos por nombre (normalizado)
    grupos_por_nombre = {g.nombre_grupo.strip().lower(): g for g in grupos}
    # índice de pruebas por nombre
    pruebas_por_nombre = {p.nombre.strip().lower(): p for p in pruebas}

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {exc}")

    # Un nuevo upload del mismo periodo debe reemplazar la captura anterior.
    db.query(ResultadoPrueba).filter(ResultadoPrueba.periodo_id == periodo_id).delete()

    total_procesadas = 0
    total_guardadas = 0
    total_saltadas = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grupo = grupos_por_nombre.get(sheet_name.strip().lower())
        if grupo is None:
            logger.warning("Hoja '%s' no corresponde a ningún grupo del seguimiento — ignorada.", sheet_name)
            continue

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            logger.warning("Hoja '%s' sin suficientes filas — ignorada.", sheet_name)
            continue

        # Fila 1 (índice 1) = encabezados
        header_row = [str(c).strip() if c is not None else "" for c in rows[1]]

        # Mapear columna → prueba_id
        col_prueba: dict[int, PruebaFisica] = {}
        for col_idx, col_name in enumerate(header_row):
            prueba = pruebas_por_nombre.get(col_name.lower())
            if prueba:
                col_prueba[col_idx] = prueba

        # Índices de columnas fijas
        try:
            idx_matricula = header_row.index("Matricula")
            idx_nombre = header_row.index("Nombre")
        except ValueError:
            logger.warning("Hoja '%s' no tiene columnas Matricula/Nombre — ignorada.", sheet_name)
            continue

        # Filas de datos (a partir de fila 2, índice 2)
        for row_idx, row in enumerate(rows[2:], start=3):
            total_procesadas += 1

            matricula = _str_val(row, idx_matricula)
            if not matricula:
                logger.info("Fila %d hoja '%s': matrícula vacía — saltada.", row_idx, sheet_name)
                total_saltadas += 1
                continue

            nombre_alumno = _str_val(row, idx_nombre)

            for col_idx, prueba in col_prueba.items():
                valor = _numeric_val(row, col_idx)

                resultado = ResultadoPrueba(
                    periodo_id=periodo_id,
                    prueba_id=prueba.id,
                    grupo_id=grupo.id,
                    matricula=matricula,
                    nombre_alumno=nombre_alumno,
                    valor=valor,
                )
                db.add(resultado)

            total_guardadas += 1

    db.commit()

    return {
        "total_procesadas": total_procesadas,
        "total_guardadas": total_guardadas,
        "total_saltadas": total_saltadas,
    }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _str_val(row, idx) -> Optional[str]:
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _int_val(row, idx) -> Optional[int]:
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _numeric_val(row, idx) -> Optional[float]:
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        logger.info("Valor no numérico en columna %d — guardado como NULL.", idx)
        return None
