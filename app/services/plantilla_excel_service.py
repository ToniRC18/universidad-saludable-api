import io
from datetime import timedelta, date
from typing import Optional
from sqlalchemy.orm import Session
from fastapi import HTTPException
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment

from app.models.semestres import Carrera, Semestre, Horario, GrupoSemestre

DAYS_MAP = {
    "lunes": 0,
    "martes": 1,
    "miercoles": 2,
    "miércoles": 2,
    "jueves": 3,
    "viernes": 4,
    "sabado": 5,
    "sábado": 5,
    "domingo": 6
}

def clean_filename(name: str) -> str:
    # Retorna un string seguro para nombre de archivo
    valid_chars = "-_.() abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    cleaned = ''.join(c for c in name if c in valid_chars)
    return cleaned.replace(" ", "_")

def _fechas_para_dias(semestre, dias_nombres: list[str]) -> list[date]:
    """Retorna fechas ordenadas del semestre que caen en los días indicados."""
    target_days = [DAYS_MAP[d.lower().strip()] 
                   for d in dias_nombres if d and d.lower().strip() in DAYS_MAP]
    if not target_days:
        return []
    fechas = []
    current = semestre.fecha_inicio
    while current <= semestre.fecha_fin:
        if current.weekday() in target_days:
            fechas.append(current)
        current += timedelta(days=1)
    return fechas

def _fechas_grupo(horario, semestre, grupo) -> list[date]:
    """Calcula las fechas específicas para un grupo según su sub-bloque."""
    if grupo.sub_bloque == 2:
        dias = [horario.dia_3, horario.dia_4]
    elif grupo.sub_bloque == 1:
        dias = [horario.dia_1, horario.dia_2]
    else:
        # Sin sub_bloque: usar todos los días no nulos del horario
        dias = [horario.dia_1, horario.dia_2, horario.dia_3, horario.dia_4]
    
    fechas = _fechas_para_dias(semestre, [d for d in dias if d])
    
    if not fechas:
        # Fallback histórico si no hay días configurados o mapeados
        return _fechas_para_dias(semestre, ["lunes", "miercoles"])
    
    return fechas

def generar_plantilla(
    db: Session, 
    semestre_id: int, 
    horario_id: int,
    sub_bloque: Optional[int] = None
) -> tuple[bytes, str]:
    semestre = db.query(Semestre).filter(Semestre.id == semestre_id).first()
    if not semestre:
        raise HTTPException(status_code=404, detail="Semestre no encontrado.")
    
    horario = db.query(Horario).filter(Horario.id == horario_id, Horario.semestre_id == semestre_id).first()
    if not horario:
        raise HTTPException(status_code=404, detail="Horario no encontrado en el semestre actual.")
    
    # Filtrado de grupos según sub_bloque
    query = db.query(GrupoSemestre).filter(GrupoSemestre.horario_id == horario_id)
    if sub_bloque is not None:
        query = query.filter(GrupoSemestre.sub_bloque == sub_bloque)
    
    grupos = query.all()
    if not grupos:
        detail = "Este horario no tiene grupos"
        if sub_bloque: detail += f" en el sub-bloque {sub_bloque}"
        raise HTTPException(status_code=400, detail=detail)

    carreras = db.query(Carrera).filter(Carrera.activa == True).order_by(Carrera.nombre).all()

    wb = openpyxl.Workbook()
    # Hoja de catálogo oculta
    ws_cat = wb.active
    ws_cat.title = "_catalogo"
    for i, c in enumerate(carreras, start=1):
        ws_cat.cell(row=i, column=1, value=c.nombre)
    ws_cat.sheet_state = 'hidden'

    # Hoja de metadatos oculta
    from datetime import datetime
    ws_meta = wb.create_sheet(title="_meta")
    ws_meta.cell(row=1, column=1, value="semestre_id")
    ws_meta.cell(row=1, column=2, value=semestre_id)
    ws_meta.cell(row=2, column=1, value="horario_id")
    ws_meta.cell(row=2, column=2, value=horario_id)
    ws_meta.cell(row=3, column=1, value="generada_at")
    ws_meta.cell(row=3, column=2, value=datetime.now().isoformat())
    ws_meta.sheet_state = 'hidden'

    # Crea un sheet por cada grupo
    for idx, grupo in enumerate(grupos):
        ws = wb.create_sheet(title=grupo.nombre[:31]) # Excel limits to 31 chars
        
        # Obtener fechas específicas para este grupo
        fechas_sesion = _fechas_grupo(horario, semestre, grupo)
        
        # Fila 1: Encabezado general
        header_text = f"{semestre.nombre} | {horario.nombre} | Valor por sesión: {semestre.valor_por_sesion}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(fechas_sesion))
        cell_h1 = ws.cell(row=1, column=1, value=header_text)
        cell_h1.font = Font(bold=True, size=12)
        cell_h1.alignment = Alignment(horizontal='center', vertical='center')

        # Fila 2: Headers de columnas
        headers = ["Nombre", "Matrícula", "Carrera"]
        # Agregamos las fechas en formato DD/MM/YYYY
        headers.extend([f.strftime("%d/%m/%Y") for f in fechas_sesion])

        for col_idx, header_name in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Ajuste de ancho de columnas
        ws.column_dimensions['A'].width = 35 # Nombre
        ws.column_dimensions['B'].width = 15 # Matrícula
        ws.column_dimensions['C'].width = 45 # Carrera
        
        for col_idx in range(4, 4 + len(fechas_sesion)):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12

        # Congelar primera fila de headers (hasta la fila 2 que son las cabeceras)
        ws.freeze_panes = ws['A3']

        # Filas 3 a 12 (prellenado vacío)
        if carreras:
            list_formula = f'_catalogo!$A$1:$A${len(carreras)}'
            dv = DataValidation(type="list", formula1=list_formula, allow_blank=True)
            # Aplicar a C3:C12
            dv.add(f'C3:C12')
            ws.add_data_validation(dv)

    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    
    # Nombre de archivo dinámico
    mode_suffix = "completa"
    if sub_bloque == 1: mode_suffix = "sub-bloque1"
    if sub_bloque == 2: mode_suffix = "sub-bloque2"
    
    filename = f"plantilla_{clean_filename(horario.nombre)}_{mode_suffix}.xlsx"
    return file_bytes.getvalue(), filename
